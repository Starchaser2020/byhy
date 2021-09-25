import requests
import openpyxl

params = {
    'username': 'byhy',
    'password': '88888888'
}
s = requests.Session()

s.post('http://127.0.0.1:8047/api/mgr/signin', data=params)

customers =s.get('http://127.0.0.1:8047/api/mgr/customers',
           params={
               "action":"list_customer",
               "pagesize":100,
               "pagenum":1,
           }).json()["retlist"]
print(s.get('http://127.0.0.1:8047/api/mgr/customers',
           params={
               "action":"list_customer",
               "pagesize":100,
               "pagenum":1,
           }))
title = list(customers[0].keys())
book = openpyxl.Workbook()
sh = book.active
sh.title = "客户"
sh.append(title)
for customer in customers:
    info = [customer[title[0]],customer[title[1]],customer[title[2]],customer[title[3]]]
    sh.append(info)




medicines =s.get('http://127.0.0.1:8047/api/mgr/medicines',
           params={
               "action":"list_medicine",
               "pagesize":100,
               "pagenum":1,
           }).json()["retlist"]

title = list(medicines[0].keys())
sh2 = book.create_sheet('药品')

sh2.append(title)
for medicine in medicines:
    info = [medicine[title[0]],medicine[title[1]],medicine[title[2]],medicine[title[3]]]
    sh2.append(info)

orders =s.get('http://127.0.0.1:8047/api/mgr/orders',
           params={
               "action":"list_order",
               "pagesize":100,
               "pagenum":1,
           }).json()["retlist"]

title = list(orders[0].keys())
sh3 = book.create_sheet('订单')
print(title)
sh3.append(title)
for order in orders:
    info = [order[title[0]],order[title[1]],order[title[2]],order[title[3]],order[title[4]]]
    sh3.append(info)





book.save('信息.xls')
