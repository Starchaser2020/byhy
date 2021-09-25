import json
import os
import openpyxl
book = openpyxl.Workbook()
sh = book.active
sh.title = '引航'
egNames = ['startWorkTime','shipNameCn','shipNameEn','shipFlag','shipLength','max_water_draft','dynamicName','startBerth','endBerth','master',
            'assistant','assistant2','orgShort','dTelephone','channelName','headThruster','tailThruster','remarks']
cnNames = ['计划时间','中文船名','英文船名','国籍(地区)' ,'船长','	吃水','	动态','	起点泊位','	终点泊位','主引',
           '副引','其它引水','代理','代理电话','航道','头侧推','尾侧推','备注']

sh.append(cnNames)
with open('info1.txt','r',encoding='utf8') as f:
    lines = f.readlines()
    for line in lines:
        jsonLine = json.loads(line)
        for info in jsonLine:
            singleShip = []
            for i in range(len(cnNames)):
                if egNames[i] not in info:
                    item = ''
                else:
                    item = info[egNames[i]]
                singleShip.append(item)
            sh.append(singleShip)
sh.cell(1,16).value = '侧推'
for i in range(2,sh.max_row+1):
    sh.cell(i,16).value = f'首：{sh.cell(i,16).value} 尾：{sh.cell(i,17).value}'
sh.delete_cols(17)
book.save('引航表1.xls')
os.startfile('引航表1.xls')
